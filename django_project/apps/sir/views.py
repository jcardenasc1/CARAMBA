# -*- encoding: utf-8 -*-
from django.shortcuts import render
from django.views.generic import CreateView,UpdateView,DeleteView, TemplateView, View,FormView
from django.core.urlresolvers import reverse_lazy
from apps.sir.models import Exercise,Evaluation,Characteristic,Semester,Course,registration
from apps.usuario.models import User
from django.http import HttpResponse,JsonResponse
from django.http import HttpResponseRedirect
from .forms import EvaluationForm, CharacteristicForm,ExerciseForm,SemesterForm,CourseForm
# Create your views here.
import random
from datetime import datetime, date, time, timedelta
from apps.sir.htmltopdf import render_to_pdf
from openpyxl import Workbook
from django.contrib import messages

# CRUD Semestre
class listSemester(TemplateView):
    template_name = 'sir/panelSemestre.html'

    def get_context_data(self, **kwargs):
        context=super(listSemester, self).get_context_data(**kwargs)
        semester = Semester.objects.all().order_by('-created')

        context['semestre']= semester

        return context


class newSemester(CreateView):
    template_name = 'sir/crearSemestre.html'
    model=Semester
    form_class = SemesterForm
    success_url = reverse_lazy('adminSemester')

    # def form_valid(self,form):
    #     form.instance.user = self.request.user
    #
    #     return super(newSemester, self).form_valid(form)


class editSemester(UpdateView):
    template_name = 'sir/crearSemestre.html'
    model=Semester
    form_class = SemesterForm
    success_url = reverse_lazy('adminSemester')

#/ Crud Exercise


# CRUD Course
class listCourse(TemplateView):
    template_name = 'sir/panelCurso.html'

    def get_context_data(self, **kwargs):
        context=super(listCourse, self).get_context_data(**kwargs)
        course = Course.objects.all().order_by('-created')

        context['curso']= course

        return context


class newCourse(CreateView):
    template_name = 'sir/crearCurso.html'
    model=Course
    form_class = CourseForm
    success_url = reverse_lazy('adminCourse')

    # def form_valid(self,form):
    #     form.instance.user = self.request.user
    #
    #     return super(newSemester, self).form_valid(form)


class editCourse(UpdateView):
    template_name = 'sir/crearCurso.html'
    model=Course
    form_class = CourseForm
    success_url = reverse_lazy('adminCourse')

#/ Crud Exercise


# CRUD Exercise
class listExercise(TemplateView):
    template_name = 'usuario/panelEjercicios.html'

    def get_context_data(self, **kwargs):
        context=super(listExercise, self).get_context_data(**kwargs)
        if self.request.user.is_superuser:
            ejer = Exercise.objects.all().order_by('-created')
        else:
            ejer = Exercise.objects.filter(user= self.request.user).order_by('-created')

        context['ejercicio']= ejer

        return context


class newExercise(CreateView):
    template_name = 'usuario/crearejercicios.html'
    model=Exercise
    form_class = ExerciseForm
    success_url = reverse_lazy('adminExercise')

    def form_valid(self,form):
        form.instance.user = self.request.user
        messages.success(self.request, 'Guardado sastifactoriamente!')
        return super(newExercise, self).form_valid(form)


class editExercise(UpdateView):
    template_name = 'usuario/crearejercicios.html'
    model=Exercise
    form_class = ExerciseForm
    success_url = reverse_lazy('adminExercise')

    def form_valid(self,form):
        form.instance.user = self.request.user
        messages.success(self.request, 'Guardado sastifactoriamente!')
        return super(editExercise, self).form_valid(form)

#/ Crud Exercise
#Crud characteristic
class listCharacteristic(TemplateView):
    template_name = 'usuario/panelCaracteristica.html'

    def get_context_data(self, **kwargs):
        context=super(listCharacteristic, self).get_context_data(**kwargs)
        caracteristica = Characteristic.objects.all().order_by('-created')

        context['caracteristica']= caracteristica

        return context


class newCharacteristic(CreateView):
    template_name = 'usuario/crearCaracteristica.html'
    model=Characteristic
    form_class = CharacteristicForm
    success_url = reverse_lazy('adminCharacteristic')

    # def form_valid(self,form):
    #     form.instance.user = self.request.user
    #
    #     return super(newCharacteristic, self).form_valid(form)


class editCharacteristic(UpdateView):
    template_name = 'usuario/crearCaracteristica.html'
    model=Characteristic
    form_class = CharacteristicForm
    success_url = reverse_lazy('adminCharacteristic')

#/Crud characteristic


# Mis ejercicios resueltos
class searchExercise(TemplateView):
    template_name = 'usuario/ejercicios.html'

    def get_context_data(self, **kwargs):
        context=super(searchExercise, self).get_context_data(**kwargs)

        exercise_list = Evaluation.objects.filter(user = self.request.user).order_by('-created')
        context['ejercicio']=exercise_list
        context['cantidad']=context['ejercicio'].count()

        return context

#comenzar ejercicio
class startExercise(CreateView):
    template_name = 'sir/comenzar.html'
    model=Evaluation
    form_class = EvaluationForm
    success_url = reverse_lazy('start')

    def get_context_data(self, **kwargs):
        context=super(startExercise, self).get_context_data(**kwargs)
        edad = int((datetime.now().date() - self.request.user.birthdate).days / 365.25)
        if not Evaluation.objects.filter(user=self.request.user).exists() and edad<=13:
            misEjercicios = Exercise.objects.filter(state=True, range=1).order_by('level')
        else:
            misEjercicios = Exercise.objects.all().exclude(evaluation__user=self.request.user)
        # print misEjercicios.exists()
        if misEjercicios.exists():
            for e in misEjercicios:
                ejer = Exercise.objects.get(pk=e.id)
                context['ejercicio']= ejer
        else:
            context['ejercicio']= "no hay"

        return context

    def form_valid(self,form):
        form.instance.user = self.request.user
        # form.instance.exercise = Exercise.objects.get(pk=1)
        form.instance.score=random.randint(5,21)
        form.instance.abstraction=aleatorio()
        form.instance.parallelization=aleatorio()
        form.instance.logic=aleatorio()
        form.instance.synchronization=aleatorio()
        form.instance.flowControl=aleatorio()
        form.instance.userInteractivity=aleatorio()
        form.instance.dataRepresentation=aleatorio()
        print random.randint(1, 5)
        messages.success(self.request, 'Guardado sastifactoriamente!')
        return super(startExercise, self).form_valid(form)

def aleatorio():
    a = random.randint(0, 3)
    return a

#todos los ejercicios
class allExercise(TemplateView):

    def get(self,request,*args,**kwargs):
        edad = int((datetime.now().date() - self.request.user.birthdate).days / 365.25)
        if edad >= 17:
            ejercicio = Exercise.objects.filter(state=True,range=3).order_by('created')
        else:
            if edad <=13:
                ejercicio = Exercise.objects.filter(state=True, range=1).order_by('created')

        ahora = datetime.now()
        nuevo=""
        context=[]
        for e in ejercicio:
            # print e.created.month, e.created.day,e.created.year
            # print ahora.month,ahora.day,ahora.year
            if e.created.year == ahora.year and e.created.month == ahora.month and e.created.day <= ahora.day:
                nuevo=" <i class='fa fa-star' aria-hidden='true'><b>Nuevo</b></i>"
            else:
                nuevo=""

            str = e.user.first_name + ' ' + e.user.last_name
            json = {
                        "id": e.id,
                        "usuario": str.upper(),
                        "nombre": e.name + nuevo,
                        "descripcion": e.description,
                        "creado": e.created.strftime("%Y/%m/%d a las %H:%M:%S"),
                    }
            context.append(json)

        return JsonResponse(context,safe=False)


#respuesta json de ejercicio seleccionado
class Search(TemplateView):

    def get(self,request,*args,**kwargs):
        # print request.GET['id']
        ejercicio = Exercise.objects.get(pk=request.GET['id'])
        context=[]
        str = ejercicio.user.first_name + ' ' + ejercicio.user.last_name
        json = {
                    "id": ejercicio.id,
                    "usuario": str.upper(),
                    "nombre": ejercicio.name,
                    "descripcion": ejercicio.description,
                    "creado": ejercicio.created.strftime("%d/%m/%Y a las %H:%M:%S"),
                }
        context.append(json)

        return JsonResponse(context,safe=False)

# Recomendacion de ejercicios
class Recommended(TemplateView):
    def get(self,request,*args,**kwargs):
        conta=0.0
        ejercicios = Evaluation.objects.filter(user=self.request.user)

        # array de mis ejercicios
        misEjercicios = Evaluation.objects.values_list('exercise__pk', flat=True).filter(user=self.request.user)
        # elimino duplicados
        listaEjercicios = list(set(misEjercicios))
        # usuarios que han resuelto mis ejercicios
        usuarios = Evaluation.objects.values_list('user__pk', flat=True).filter(exercise__pk__in=misEjercicios).exclude(user=self.request.user)
        # eliminar usuarios duplicados
        usuarioc = list(set(usuarios))

        # print misEjercicios,'mis ejercicios'
        # print usuarios,'lista de usuarios'
        # print len(usuarios),'cantidad de usuario'
        # print Evaluation.objects.filter(user__pk=17).count()
        # for e in Evaluation.objects.filter(user__pk=17):
        #     print e.exercise.pk

        context=[]
        conta = len(misEjercicios)
        rango = 4.0

        calGusto = 0.0
        calDificultad = 0.0
        promedio = 0.0
        puntuacion=[]
        # usuarioc=[]
        usuar=[]
        # listaEjercicios=[]
        # sacar listado de mis ejercicios resueltos y sacar el listado de todos los usuarios q han resuelto un ejercicio
        # for e in ejercicios:
        #     usuarios = Evaluation.objects.filter(exercise=e.exercise).exclude(user=self.request.user)
        #     listaEjercicios.append(e.exercise.id)
        #     for comun in usuarios:
        #         if comun.user.id not in usuarioc:
        #             usuarioc.append(comun.user.id)
        #
        c=0.0
        for x in usuarioc:
                ejerciciosExiste = Evaluation.objects.filter(user__id=x, exercise__pk__in=misEjercicios)
                for ejerc in ejercicios:
                    for ec in ejerciciosExiste:
            #             # corrigio get por filter en ejercicioComun
            #             ejerciciosComun = Evaluation.objects.filter(user__id=x, exercise=ejerc.exercise)
            #             for ec in ejerciciosComun:

                        gusto = ((ejerc.like - ec.like) / rango) ** 2
                        calGusto = calGusto + gusto
                        dificultad = ((ejerc.difficulty - ec.difficulty) / rango) ** 2
                        calDificultad = calDificultad + dificultad
                        c += 1
        #
        #             # antes de la correccion
        #             # ejerciciosComun = Evaluation.objects.get(user__id=x, exercise=ejerc.exercise)
        #             # gusto = ((ejerc.like-ejerciciosComun.like)/rango) ** 2
        #             # calGusto = calGusto + gusto
        #             # dificultad = ((ejerc.difficulty-ejerciciosComun.difficulty)/rango) ** 2
        #             # calDificultad = calDificultad + dificultad
        #             # c+=1
        #     # promedio = ((1-((1/c)*calGusto))+(1-((1/c)*calDificultad)))/2
                promedio = ((((1-((1/c)*calGusto))+(1-((1/c)*calDificultad)))/2)+(c/conta))/2
                puntuacion.append(promedio)
                c=0.0
                calGusto = 0.0
                calDificultad = 0.0
        #
        #
        orden = sorted(puntuacion,reverse = True)
        for y in range(len(orden)):
            for z in range(len(puntuacion)):
                if orden[y] == puntuacion[z]:
                    usuar.append(usuarioc[z])
        print len(usuarioc)
        print len(list(set(usuar)))
        usuar = list(set(usuar))
        #
        recomendacion = []
        # for idUsuario in usuar:
        # for eje in ejercicios:
        recomienda = Evaluation.objects.values_list('exercise__pk', flat=True).filter(user__id__in=usuar[:5]).exclude(exercise__id__in=misEjercicios)
        # print len(recomienda), 'cantidad'
        recomendacion = list(set(recomienda))

        print len(recomendacion)
        # misEjercicios = Evaluation.objects.values_list('exercise__pk', flat=True).filter(user=self.request.user)
        # elimino duplicados
        # listaEjercicios = list(set(misEjercicios))

        # 1411
        # for r in recomienda:
        #     # print r.exercise.name
        #     if r.exercise.id not in recomendacion:
        #         recomendacion.append(r.exercise.id)
        for idRecomendado in recomendacion:
            ejeRecomendado=Exercise.objects.get(pk=idRecomendado)
            str = ejeRecomendado.user.first_name + ' ' + ejeRecomendado.user.last_name
            json = {
                    "id": ejeRecomendado.id,
                    "usuario": str.upper(),
                    "nombre": ejeRecomendado.name,
                    "descripcion": ejeRecomendado.description,
                    "creado": ejeRecomendado.created.strftime("%d/%m/%Y a las %H:%M:%S"),
                }
            context.append(json)


        return JsonResponse(context,safe=False)

# Nivel Academico del estudiante
class Level(TemplateView):

    def get(self,request,*args,**kwargs):
        ejercicio = Evaluation.objects.filter(user=self.request.user)
        total=ejercicio.count()
        context=[]
        abstraction=0
        parallelization=0
        logic=0
        synchronization=0
        flowControl=0
        userInteractivity=0
        dataRepresentation=0

        for ejer in ejercicio:
            abstraction+=ejer.abstraction
            parallelization+=ejer.parallelization
            logic+=ejer.logic
            synchronization+=ejer.synchronization
            flowControl+=ejer.flowControl
            userInteractivity+=ejer.userInteractivity
            dataRepresentation+=ejer.dataRepresentation
        if ejercicio.exists():
            json = {
                        "abstraction": abstraction/total,
                        "parallelization": parallelization/total,
                        "logic": logic/total,
                        "synchronization": synchronization/total,
                        "flowControl": flowControl/total,
                        "userInteractivity": userInteractivity/total,
                        "dataRepresentation": dataRepresentation/total,
                    }
            context.append(json)

        return JsonResponse(context,safe=False)


class myStudent(TemplateView):
    def get(self,request,*args,**kwargs):

        context=[]
        students_list = User.objects.filter(registration__course__user=self.request.user).order_by('last_name')

        for student in students_list:
            curso = registration.objects.get(user=student)
            ejercicios=Evaluation.objects.filter(user=student)
            str = student.first_name + ' ' + student.last_name

            json = {
                        "id_estudiante": student.id,
                        "estudiante": str,
                        "curso": curso.course.semester.description + " - " + curso.course.description ,
                        "cantidad": ejercicios.count(),
                        "paralelismo": promedio(student,ejercicios.count(),0),
                        "logica": promedio(student,ejercicios.count(),1),
                        "abstraccion": promedio(student,ejercicios.count(),2),
                        "interactua": promedio(student,ejercicios.count(),3),
                        "control": promedio(student,ejercicios.count(),4),
                        "representa": promedio(student,ejercicios.count(),5),
                        "sincroniza": promedio(student,ejercicios.count(),6),

                    }
            context.append(json)

        return JsonResponse(context,safe=False)


def promedio(user, cant,c):

    ejercicios=Evaluation.objects.filter(user=user)
    a=0
    p=0
    l=0
    s=0
    f=0
    u=0
    d=0
    if cant == 0:
        cant = 1
    for e in ejercicios:
        a =a + e.abstraction
        p =p + e.parallelization
        l =l + e.logic
        s =s + e.synchronization
        f =f + e.flowControl
        u =u + e.userInteractivity
        d =d + e.dataRepresentation

    a =str((a/cant)*100/3)+"%"
    p =str((p/cant)*100/3)+"%"
    l =str((l/cant)*100/3)+"%"
    s =str((s/cant)*100/3)+"%"
    f =str((f/cant)*100/3)+"%"
    u =str((u/cant)*100/3)+"%"
    d =str((d/cant)*100/3)+"%"

    carac=[p,l,a,u,f,d,s]

    return carac[c]


class showExercise(TemplateView):
    def get(self,request,*args,**kwargs):

        context=[]
        exercise = Evaluation.objects.filter(user=request.GET['id']).order_by('-created')

        for e in exercise:
            str = e.user.first_name + ' ' + e.user.last_name
            json = {
                        "estudiante": str,
                        "ejercicio": "<a href="+e.file.url+">" + e.exercise.name + "</a>",
                        "duracion": e.duration,
                        "fecha": e.created.strftime("%Y/%m/%d a las %H:%M:%S"),
                        "gusto": e.like,
                        "dificultad": e.difficulty,

                    }
            context.append(json)

        return JsonResponse(context,safe=False)


#Nivel academico

def myLevel(request):
    puntaje=0
    ejer=Evaluation.objects.filter(user=request.user)
    for e in ejer:
        puntaje=puntaje + e.score

    puntaje=puntaje/ejer.count()

    context = {
               'fecha':datetime.now,
               'usuario':request.user,
               'puntaje':str((puntaje*100)/21)+"/100",

    }
    pdf = render_to_pdf('sir/nivelAcademico.html',context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Certificado_%s.pdf" %(request.user)
        content = "inline; filename='%s'" %(filename)
        download = request.GET.get("download")
        if download:
            content = "attachment; filename='%s'" %(filename)
        response['Content-Disposition'] = content
        return response
    return HttpResponse("Not found")


#Reportes excel

import xlwt
from xlwt import *

def ReportUserExcel(request):

    try:
        __author__ = 'CarambaScratch'
        style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
        style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
        style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
        title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
        style1 = easyxf(num_format_str='D-MMM-YY')
        font_style = XFStyle()
        font_style.font.bold = True
        font_style2 = XFStyle()
        font_style2.font.bold = False
        wb = Workbook(encoding='utf-8')
        ws = wb.add_sheet('Usuarios')
        # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
        response = HttpResponse(content_type='application/ms-excel')
        response[
            'Content-Disposition'] = 'attachment; filename=Reporte_de_usuarios' + random.randint(
            1, 10000).__str__() + '.xls'

        columns = [
            (u"N.", 1500),
            (u"APELLIDOS / NOMBRES", 10000),
            (u"EDAD", 3000),
            (u"SEXO", 3000),
            (u"SEMESTRE", 4000),
            (u"DOCENTE", 10000),
        ]
        row_num = 0
        for col_num in xrange(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            ws.col(col_num).width = columns[col_num][1]
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'yyyy/mm/dd'
        row_num = 1
        i = 0
        users = User.objects.all()
        ahora = datetime.now()
        for usuarios in users :
            sexo = ['M', 'F', 'O']
            campo1 = usuarios.last_name + " " + usuarios.first_name
            campo2 = ahora.year - usuarios.birthdate.year
            campo3 = sexo[int(usuarios.sex)-1]
            if registration.objects.filter(user=usuarios.id).exists():
                semestre = registration.objects.get(user=usuarios.id)
                campo4 = semestre.course.semester.description + " " +semestre.course.description
                campo5 = semestre.course.user.last_name + " " + semestre.course.user.first_name
            else:
                semestre = "No hay registro"
                campo4 = semestre
                campo5 = semestre

            i += 1
            ws.write(row_num, 0, i, font_style2)
            ws.write(row_num, 1, campo1, font_style2)
            ws.write(row_num, 2, campo2, font_style2)
            ws.write(row_num, 3, campo3, font_style2)
            ws.write(row_num, 4, campo4, font_style2)
            ws.write(row_num, 5, campo5, font_style2)
            row_num += 1
        wb.save(response)
        return response
    except Exception as ex:
        pass

def ReportExerciseExcel(request):
    try:
        __author__ = 'CarambaScratch'
        style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
        style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
        style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
        title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
        style1 = easyxf(num_format_str='D-MMM-YY')
        font_style = XFStyle()
        font_style.font.bold = True
        font_style2 = XFStyle()
        font_style2.font.bold = False
        wb = Workbook(encoding='utf-8')
        ws = wb.add_sheet('Ejercicios')
        # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
        response = HttpResponse(content_type='application/ms-excel')
        response[
            'Content-Disposition'] = 'attachment; filename=Reporte_de_ejercicios' + random.randint(
            1, 10000).__str__() + '.xls'

        columns = [
            (u"N.", 1500),
            (u"NOMBRE", 12000),
            (u"CARACTERISTICA", 5000),
            (u"DIFICULTAD", 3000),
            (u"ESTADO", 3000),
        ]
        row_num = 0

        for col_num in xrange(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            ws.col(col_num).width = columns[col_num][1]
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'yyyy/mm/dd'
        row_num = 1
        i = 0
        exercise = Exercise.objects.all()

        for ejercicios in exercise :
            sexo = ['M', 'F', 'O']
            campo1 = ejercicios.name
            campo2 = ejercicios.characteristic.description
            campo3 = ejercicios.level
            campo4 = ejercicios.state

            i += 1
            ws.write(row_num, 0, i, font_style2)
            ws.write(row_num, 1, campo1, font_style2)
            ws.write(row_num, 2, campo2, font_style2)
            ws.write(row_num, 3, campo3, font_style2)
            ws.write(row_num, 4, campo4, font_style2)
            row_num += 1
        wb.save(response)
        return response
    except Exception as ex:
        pass


def ReportEvaluationExcel(request):
    try:
        __author__ = 'CarambaScratch'
        style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
        style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
        style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
        title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
        style1 = easyxf(num_format_str='D-MMM-YY')
        font_style = XFStyle()
        font_style.font.bold = True
        font_style2 = XFStyle()
        font_style2.font.bold = False
        wb = Workbook(encoding='utf-8')
        ws = wb.add_sheet('Evaluacion')
        # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
        response = HttpResponse(content_type='application/ms-excel')
        response[
            'Content-Disposition'] = 'attachment; filename=Evaluacion' + random.randint(
            1, 10000).__str__() + '.xls'

        columns = [
            (u"N.", 1500),
            (u"NOMBRE", 4000),
            (u"ID", 500),
            (u"CARACTERISTICA", 4000),
            (u"NIVEL", 3000),
            (u"DURACION", 3000),
            (u"GUSTO", 3000),
            (u"DIFICULTAD", 4000),
            (u"ESTUDIANTE", 12000),
        ]
        row_num = 0
        for col_num in xrange(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            ws.col(col_num).width = columns[col_num][1]
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'yyyy/mm/dd'
        row_num = 1
        i = 0
        x = Evaluation.objects.all().order_by('user','id')
        for evaluacion in x:
            campo1 = evaluacion.exercise.name
            campo2 = evaluacion.exercise.id
            campo3 = evaluacion.exercise.characteristic.description
            campo4 = evaluacion.exercise.level
            campo5 = evaluacion.duration
            campo6 = evaluacion.like
            campo7 = evaluacion.difficulty
            campo8 = evaluacion.user.last_name + " " + evaluacion.user.first_name
            i += 1
            ws.write(row_num, 0, i, font_style2)
            ws.write(row_num, 1, campo1, font_style2)
            ws.write(row_num, 2, campo2, font_style2)
            ws.write(row_num, 3, campo3, font_style2)
            ws.write(row_num, 4, campo4, font_style2)
            ws.write(row_num, 5, campo5, font_style2)
            ws.write(row_num, 6, campo6, font_style2)
            ws.write(row_num, 7, campo7, font_style2)
            ws.write(row_num, 8, campo8, font_style2)
            row_num += 1
        wb.save(response)
        return response
    except Exception as ex:
        pass
